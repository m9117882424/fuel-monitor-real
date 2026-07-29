#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel: пробег группы Arvento + заправки Turpak/Shell/Petrol.

По умолчанию Arvento запрашивается блоками по 14 дней через
VehicleDistanceReport2. Режимы: 14d, month, daily. Если блок не возвращает
датированную детализацию, только этот блок повторяется посуточно через
VehicleDistanceReport.
"""
from __future__ import annotations

import argparse
import getpass
import os
import re
import time
import unicodedata
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date, datetime, time as dt_time, timedelta
from pathlib import Path
from typing import Any, Iterable, Sequence

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import inspect, text
from sqlalchemy.engine import Engine

try:
    from app.db import SessionLocal, engine
except ImportError as exc:
    raise SystemExit(
        "Запускайте файл из корня fuel-monitor-real с активированным venv."
    ) from exc

ARVENTO_URL = "https://ws.arvento.com/v1/report.asmx"
DATE_FMT = "%Y%m%d%H%M%S"
TZ = "Europe/Istanbul"
SOURCES = ("Turpak", "Shell", "Petrol")
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/126 Safari/537.36",
    "Accept": "text/xml,application/xml,text/html,*/*",
    "Origin": "https://ws.arvento.com",
}


@dataclass(frozen=True)
class Credentials:
    username: str
    pin1: str
    pin2: str


def norm_plate(value: Any) -> str:
    return re.sub(r"[^0-9A-ZА-ЯЁ]", "", str(value or "").strip().upper())


def number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)
    raw = re.sub(r"[^0-9,.-]", "", str(value).replace("\u00a0", "").replace(" ", ""))
    if not raw or raw in {"-", ".", ","}:
        return 0.0
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".") if raw.rfind(",") > raw.rfind(".") else raw.replace(",", "")
    else:
        raw = raw.replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def xml_key(name: str) -> str:
    name = name.split("}")[-1]
    name = re.sub(r"_x([0-9A-Fa-f]{4})_", lambda m: chr(int(m.group(1), 16)), name)
    name = unicodedata.normalize("NFKD", name)
    name = "".join(ch for ch in name if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]", "", name.lower())


def first(row: dict[str, Any], aliases: Sequence[str], default: Any = "") -> Any:
    for alias in aliases:
        value = row.get(xml_key(alias))
        if value not in (None, ""):
            return value
    return default


def local_dt(value: Any) -> pd.Timestamp | pd.NaT:
    if value in (None, ""):
        return pd.NaT
    try:
        ts = pd.Timestamp(value)
    except (TypeError, ValueError):
        return pd.NaT
    if ts.tzinfo is not None:
        try:
            ts = ts.tz_convert(TZ).tz_localize(None)
        except (TypeError, ValueError):
            ts = ts.tz_localize(None)
    return ts


def dataset_rows(xml_text: str) -> list[dict[str, str]]:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError as exc:
        raise RuntimeError(f"Arvento вернул некорректный XML: {xml_text[:300]}") from exc
    errors = [(e.text or "").strip() for e in root.iter() if xml_key(e.tag) in {"error", "errormessage", "exception", "e"} and (e.text or "").strip()]
    if errors:
        raise RuntimeError("Ошибка Arvento: " + "; ".join(dict.fromkeys(errors)))
    result, seen = [], set()
    for elem in root.iter():
        row = {xml_key(c.tag): (c.text or "").strip() for c in list(elem) if not list(c)}
        if len(row) < 2:
            continue
        signature = tuple(sorted(row.items()))
        if signature not in seen:
            seen.add(signature)
            result.append(row)
    return result


def arvento_post(session: requests.Session, method: str, payload: dict[str, Any], timeout: int = 180) -> list[dict[str, str]]:
    headers = {**HEADERS, "Referer": f"{ARVENTO_URL}?op={method}"}
    last: Exception | None = None
    for attempt in range(1, 4):
        try:
            response = session.post(f"{ARVENTO_URL}/{method}", data=payload, headers=headers, timeout=timeout)
            response.raise_for_status()
            return dataset_rows(response.text)
        except (requests.RequestException, RuntimeError) as exc:
            last = exc
            if attempt < 3:
                time.sleep(attempt * 2)
    raise RuntimeError(f"Не удалось получить Arvento/{method}: {last}")


def get_vehicles(session: requests.Session, cred: Credentials, group: str) -> pd.DataFrame:
    rows = arvento_post(session, "GetNodes", {"Username": cred.username, "PIN1": cred.pin1, "PIN2": cred.pin2, "Group": group}, 90)
    items = []
    for row in rows:
        raw = first(row, ["LicensePlate", "Plaka"])
        plate = norm_plate(raw)
        if plate:
            items.append({
                "plate": plate,
                "display": str(raw).strip().upper() or plate,
                "driver": str(first(row, ["Driver", "Surucu", "Sürücü"])).strip(),
                "node": str(first(row, ["Node", "Device No", "Cihaz No"])).strip(),
            })
    if not items:
        raise RuntimeError(f"В группе Arvento {group!r} автомобили не найдены.")
    return pd.DataFrame(items).drop_duplicates("plate").sort_values("plate").reset_index(drop=True)


def empty_mileage() -> pd.DataFrame:
    return pd.DataFrame(columns=["date", "plate", "km", "driver", "node"])


def aggregate(records: list[dict[str, Any]]) -> pd.DataFrame:
    if not records:
        return empty_mileage()
    df = pd.DataFrame(records)
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.normalize()
    df = df[df["date"].notna() & df["plate"].astype(str).ne("")]
    if df.empty:
        return empty_mileage()
    return df.groupby(["date", "plate"], as_index=False).agg(km=("km", "sum"), driver=("driver", "first"), node=("node", "first")).sort_values(["date", "plate"])


def next_month(value: date) -> date:
    return date(value.year + (value.month == 12), 1 if value.month == 12 else value.month + 1, 1)


def chunks(start: date, end: date, mode: str) -> Iterable[tuple[date, date]]:
    current = start
    while current <= end:
        if mode == "daily":
            block_end = current
        elif mode == "month":
            block_end = min(next_month(current) - timedelta(days=1), end)
        else:
            block_end = min(current + timedelta(days=13), end)
        yield current, block_end
        current = block_end + timedelta(days=1)


def mileage_payload(cred: Credentials, group: str, start: date, end: date, detailed: bool) -> dict[str, str]:
    payload = {
        "Username": cred.username, "PIN1": cred.pin1, "PIN2": cred.pin2,
        "StartDate": datetime.combine(start, dt_time.min).strftime(DATE_FMT),
        "EndDate": datetime.combine(end, dt_time.max).replace(microsecond=0).strftime(DATE_FMT),
        "Node": "", "Group": group, "Compress": "", "Locale": "tr", "Language": "1",
    }
    if detailed:
        payload.update({"showSCDriver": "", "Detailed": "1"})
    return payload


def detailed_block(session: requests.Session, cred: Credentials, group: str, start: date, end: date) -> pd.DataFrame:
    rows = arvento_post(session, "VehicleDistanceReport2", mileage_payload(cred, group, start, end, True), 240)
    records = []
    for row in rows:
        plate = norm_plate(first(row, ["Plaka", "LicensePlate"]))
        event = local_dt(first(row, ["Tarih/Saat", "Tarih Saat", "Date/Time", "LocalDateTime", "Tarih", "Date"]))
        if not plate or pd.isna(event):
            continue
        day = pd.Timestamp(event).normalize()
        if start <= day.date() <= end:
            records.append({
                "date": day, "plate": plate,
                "km": max(number(first(row, ["Mesafe km", "Distance km", "Distance", "Mesafe"])), 0.0),
                "driver": str(first(row, ["Surucu", "Sürücü", "Driver"])).strip(),
                "node": str(first(row, ["Cihaz No", "Device No", "Node"])).strip(),
            })
    return aggregate(records)


def daily_block(session: requests.Session, cred: Credentials, group: str, start: date, end: date) -> pd.DataFrame:
    records, current = [], start
    while current <= end:
        print(f"  fallback: {current:%d.%m.%Y}")
        rows = arvento_post(session, "VehicleDistanceReport", mileage_payload(cred, group, current, current, False))
        for row in rows:
            plate = norm_plate(first(row, ["Plaka", "LicensePlate"]))
            if plate:
                records.append({
                    "date": current, "plate": plate,
                    "km": max(number(first(row, ["Mesafe km", "Distance km", "Distance", "Mesafe"])), 0.0),
                    "driver": str(first(row, ["Surucu", "Sürücü", "Driver"])).strip(),
                    "node": str(first(row, ["Cihaz No", "Device No", "Node"])).strip(),
                })
        current += timedelta(days=1)
    return aggregate(records)


def get_mileage(session: requests.Session, cred: Credentials, group: str, start: date, end: date, mode: str) -> pd.DataFrame:
    frames, intervals = [], list(chunks(start, end, mode))
    for i, (left, right) in enumerate(intervals, 1):
        print(f"Arvento {i}/{len(intervals)}: {left:%d.%m.%Y} — {right:%d.%m.%Y}")
        if mode == "daily":
            frame = daily_block(session, cred, group, left, right)
        else:
            try:
                frame = detailed_block(session, cred, group, left, right)
            except RuntimeError as exc:
                print(f"  Ошибка детального блока: {exc}")
                frame = empty_mileage()
            if frame.empty:
                print("  Нет датированной детализации; повторяю только этот блок посуточно.")
                frame = daily_block(session, cred, group, left, right)
        if not frame.empty:
            frames.append(frame)
    return aggregate(pd.concat(frames, ignore_index=True).to_dict("records")) if frames else empty_mileage()


def read_fuel_relation(db: Engine, name: str, start: datetime, end: datetime) -> pd.DataFrame:
    return pd.read_sql_query(text(f"""SELECT source,event_dt,plate,liters,station_code,station_name,group_name FROM {name} WHERE event_dt>=:s AND event_dt<:e"""), db, params={"s": start, "e": end})


def get_fuel(db: Engine, start: date, end: date, plates: set[str]) -> tuple[pd.DataFrame, str]:
    start_dt = datetime.combine(start, dt_time.min)
    end_dt = datetime.combine(end + timedelta(days=1), dt_time.min)
    inspector = inspect(db)
    tables, views = set(inspector.get_table_names()), set(inspector.get_view_names())
    frames, description = [], ""
    if "fuel_three_sources_v" in tables | views:
        frames = [read_fuel_relation(db, "fuel_three_sources_v", start_dt, end_dt)]
        description = "fuel_three_sources_v"
    else:
        if "fuel_events" not in tables:
            raise RuntimeError("Нет fuel_three_sources_v или fuel_events.")
        common = read_fuel_relation(db, "fuel_events", start_dt, end_dt)
        source = common["source"].fillna("").astype(str).str.lower()
        frames.append(common[source.isin({"shell", "shell_excel", "petrol"})])
        if "turpak_fuel_events_all" in tables:
            frames.append(read_fuel_relation(db, "turpak_fuel_events_all", start_dt, end_dt))
            description = "fuel_events + turpak_fuel_events_all"
        else:
            frames.append(common[source.str.contains("turpak", na=False)])
            description = "fuel_events"
    fuel = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    if fuel.empty:
        return pd.DataFrame(columns=["date", "plate", "source", "liters"]), description
    fuel["plate"] = fuel["plate"].apply(norm_plate)
    fuel["date"] = fuel["event_dt"].apply(local_dt)
    fuel["date"] = pd.to_datetime(fuel["date"], errors="coerce").dt.normalize()
    fuel["liters"] = fuel["liters"].apply(number)
    raw = fuel["source"].fillna("").astype(str).str.lower()
    fuel["source_name"] = ""
    fuel.loc[raw.str.contains("turpak", na=False), "source_name"] = "Turpak"
    fuel.loc[raw.str.contains("shell", na=False), "source_name"] = "Shell"
    fuel.loc[raw.str.contains("petrol", na=False), "source_name"] = "Petrol"
    fuel = fuel[fuel["plate"].isin(plates) & fuel["date"].notna() & fuel["source_name"].isin(SOURCES) & fuel["liters"].gt(0)]
    fuel = fuel.drop_duplicates(["source_name", "event_dt", "plate", "liters", "station_code", "station_name"])
    return fuel, description


def reports(vehicles: pd.DataFrame, mileage: pd.DataFrame, fuel: pd.DataFrame, start: date, end: date) -> tuple[pd.DataFrame, pd.DataFrame]:
    days = pd.DataFrame({"date": pd.date_range(start, end, freq="D"), "_": 1})
    cars = vehicles[["plate", "display", "driver"]].copy(); cars["_"] = 1
    daily = days.merge(cars, on="_").drop(columns="_")
    daily = daily.merge(mileage[["date", "plate", "km", "driver"]].rename(columns={"driver": "daily_driver"}), on=["date", "plate"], how="left")
    if fuel.empty:
        for source in SOURCES: daily[source] = 0.0
        daily["operations"] = 0
    else:
        liters = fuel.pivot_table(index=["date", "plate"], columns="source_name", values="liters", aggfunc="sum", fill_value=0).reset_index()
        count = fuel.groupby(["date", "plate"]).size().rename("operations").reset_index()
        daily = daily.merge(liters, on=["date", "plate"], how="left").merge(count, on=["date", "plate"], how="left")
        for source in SOURCES:
            if source not in daily: daily[source] = 0.0
    for col in ["km", *SOURCES, "operations"]: daily[col] = pd.to_numeric(daily[col], errors="coerce").fillna(0)
    daily["driver"] = daily["daily_driver"].fillna("").where(daily["daily_driver"].fillna("").ne(""), daily["driver"])
    daily["total"] = daily[list(SOURCES)].sum(axis=1)
    daily["consumption"] = daily["total"].div(daily["km"]).mul(100).where(daily["km"].gt(0), 0)
    daily["month"] = daily["date"].dt.strftime("%Y-%m")
    detail = daily[["date", "month", "display", "driver", "km", *SOURCES, "total", "operations", "consumption"]].copy()
    detail.columns = ["Дата", "Месяц", "Автомобиль", "Водитель", "Пробег, км", "Turpak, л", "Shell, л", "Petrol, л", "Заправки всего, л", "Количество заправок", "Расход, л/100 км"]
    summary = detail.groupby(["Месяц", "Автомобиль"], as_index=False).agg({"Пробег, км": "sum", "Turpak, л": "sum", "Shell, л": "sum", "Petrol, л": "sum", "Заправки всего, л": "sum", "Количество заправок": "sum"})
    summary["Расход, л/100 км"] = summary["Заправки всего, л"].div(summary["Пробег, км"]).mul(100).where(summary["Пробег, км"].gt(0), 0)
    return summary.sort_values(["Месяц", "Автомобиль"]), detail.sort_values(["Дата", "Автомобиль"])


def write_excel(summary: pd.DataFrame, detail: pd.DataFrame, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl", datetime_format="DD.MM.YYYY") as writer:
        summary.to_excel(writer, sheet_name="Сводка", index=False)
        detail.to_excel(writer, sheet_name="По дням", index=False)
    book = load_workbook(output)
    for ws in book.worksheets:
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78"); cell.font = Font(color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        for cells in ws.columns:
            width = min(max(len(str(c.value or "")) for c in cells) + 2, 38)
            ws.column_dimensions[get_column_letter(cells[0].column)].width = max(11, width)
    book.save(output)


def parse_date(value: str) -> date:
    try: return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc: raise argparse.ArgumentTypeError("Дата: ГГГГ-ММ-ДД") from exc


def credentials() -> Credentials:
    user = os.getenv("ARVENTO_USER") or input("Arvento Username: ").strip()
    pin1 = os.getenv("ARVENTO_PIN1") or getpass.getpass("Arvento PIN1: ")
    pin2 = os.getenv("ARVENTO_PIN2") or getpass.getpass("Arvento PIN2: ")
    if not all((user, pin1, pin2)): raise SystemExit("Не заполнены реквизиты Arvento.")
    return Credentials(user, pin1, pin2)


def main() -> int:
    parser = argparse.ArgumentParser(description="Пробег Arvento + Turpak/Shell/Petrol в Excel")
    parser.add_argument("--date-from", type=parse_date)
    parser.add_argument("--date-to", type=parse_date)
    parser.add_argument("--group")
    parser.add_argument("--output")
    parser.add_argument("--arvento-chunk-mode", choices=("14d", "month", "daily"), default=os.getenv("ARVENTO_CHUNK_MODE", "14d"))
    parser.add_argument("--sync-fuel", action="store_true")
    args = parser.parse_args()
    today = date.today(); start = args.date_from or today.replace(day=1); end = args.date_to or today
    if end < start: parser.error("--date-to раньше --date-from")
    group = (args.group or os.getenv("ARVENTO_GROUP") or input("Группа Arvento: ")).strip()
    if not group: parser.error("Не указана группа")
    output = Path(args.output or f"mileage_fuel_{group}_{start:%Y%m%d}_{end:%Y%m%d}.xlsx").resolve()
    cred, session = credentials(), requests.Session()
    print(f"Группа: {group}; период: {start:%d.%m.%Y}—{end:%d.%m.%Y}; режим: {args.arvento_chunk_mode}")
    vehicles = get_vehicles(session, cred, group); plates = set(vehicles["plate"])
    mileage = get_mileage(session, cred, group, start, end, args.arvento_chunk_mode)
    if args.sync_fuel:
        from app.services.sync_service import sync_all
        db = SessionLocal()
        try: sync_all(db, build_report=False, send_report=False)
        finally: db.close()
    fuel, source = get_fuel(engine, start, end, plates)
    summary, detail = reports(vehicles, mileage, fuel, start, end)
    write_excel(summary, detail, output)
    print(f"Автомобилей: {len(vehicles)}; источник топлива: {source}; Excel: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
