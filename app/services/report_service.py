from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
WARNING_FILL = PatternFill('solid', fgColor='FFF2CC')
CRITICAL_FILL = PatternFill('solid', fgColor='FCE4D6')
EXCEEDED_FILL = PatternFill('solid', fgColor='F4CCCC')
UNLIMITED_FILL = PatternFill('solid', fgColor='DBEAFE')


def _auto_fit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)) + 2)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 12), 40)


def _write_df(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in df.itertuples(index=False):
        ws.append(list(row))
    ws.freeze_panes = 'A2'
    _auto_fit(ws)


def _highlight_by_status(ws) -> None:
    headers = [c.value for c in ws[1]]
    if 'status' not in headers:
        return
    idx = headers.index('status') + 1
    for row_idx in range(2, ws.max_row + 1):
        status = ws.cell(row=row_idx, column=idx).value
        fill = None
        if status == 'WARNING':
            fill = WARNING_FILL
        elif status == 'CRITICAL':
            fill = CRITICAL_FILL
        elif status == 'EXCEEDED':
            fill = EXCEEDED_FILL
        elif status == 'UNLIMITED':
            fill = UNLIMITED_FILL
        if fill:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill


def _prepare_alert_registry(alert_registry: pd.DataFrame) -> pd.DataFrame:
    if alert_registry is None or alert_registry.empty:
        return pd.DataFrame(columns=[
            'year_month', 'plate', 'limit_bucket', 'first_threshold_pct', 'max_threshold_pct',
            'usage_pct', 'remaining_liters', 'status', 'limit_liters', 'consumed_liters',
            'total_amount_try', 'mode', 'sources', 'last_event_dt', 'first_triggered_at', 'last_seen_at',
        ])
    cols = [
        'year_month', 'plate', 'limit_bucket', 'first_threshold_pct', 'max_threshold_pct',
        'usage_pct', 'remaining_liters', 'status', 'limit_liters', 'consumed_liters',
        'total_amount_try', 'mode', 'sources', 'last_event_dt', 'first_triggered_at', 'last_seen_at',
    ]
    existing = [c for c in cols if c in alert_registry.columns]
    return alert_registry[existing].sort_values(['usage_pct', 'plate'], ascending=[False, True]).reset_index(drop=True)


def export_report(output_path: str | Path, monthly_summary: pd.DataFrame, details: pd.DataFrame, alert_registry: pd.DataFrame, driver_registry: pd.DataFrame | None = None) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet('Summary')
    _write_df(ws_summary, monthly_summary if monthly_summary is not None else pd.DataFrame())
    _highlight_by_status(ws_summary)

    ws_details = wb.create_sheet('Details')
    _write_df(ws_details, details if details is not None else pd.DataFrame())

    ws_alerts = wb.create_sheet('Alerts')
    _write_df(ws_alerts, _prepare_alert_registry(alert_registry))
    _highlight_by_status(ws_alerts)

    _strip_excel_timezones(wb)
    wb.save(output_path)
    return output_path


def _strip_excel_timezones(wb):
    """Excel/openpyxl cannot save timezone-aware datetime values."""
    import datetime as _dt

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, _dt.datetime) and value.tzinfo is not None:
                    cell.value = value.replace(tzinfo=None)
                elif isinstance(value, _dt.time) and value.tzinfo is not None:
                    cell.value = value.replace(tzinfo=None)
